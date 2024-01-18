from openpyxl import Workbook, load_workbook 
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time

terms=[]
examples=[]
seen=[]
##izveido sarakstus bez atkartosanas, bez tuksam rindam, bez tulkojumiem
with open("eb5a1db34d335268.csv",'r',errors='ignore') as file:
    for line in file:
        row= line.rstrip().split(",")
        if row[0]:
            lowercase_term= row[0].lower()
            if lowercase_term not in seen:
                terms.append(lowercase_term.capitalize())
                examples.append(row[2])
                seen.append(lowercase_term)
##izveido Excel failus no saglabata saraksta
wb=Workbook()
ws=wb.active
for i in range(len(terms)):
    ws["A"+str(i+1)].value=terms[i]
    ws["C"+str(i+1)].value=examples[i]
wb.save("result.xlsx")

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)
##atver google tulkotaju
url="https://consent.google.lv/m?continue=https://translate.google.lv/?hl%3Dlv&gl=LV&m=0&pc=t&cm=2&hl=lv&src=1"
driver.get(url)
time.sleep(2)
##uzklikskina uz anglu valodu izvelnes
find=driver.find_element(By.XPATH,"//*[@id=\"i14\"]/span[3]")
find.click()

tr=[]
##uzklikskina uz ievades lauka
find=driver.find_element(By.CLASS_NAME,"er8xn")
find.click()
##ievada terminus, saglaba tulkojumu no izvade lauka sarakstaa, iztira ievades lauku
for i in terms:
    find.send_keys(i)
    time.sleep(3)
    find=driver.find_element(By.CLASS_NAME,"ryNqvb")
    temp=find.get_attribute("textContent")
    tr.append(temp)
    find=driver.find_element(By.CLASS_NAME,"er8xn")
    find.click()
    find.clear()


## ievada tulkojumu attieciga kolonna
for i in range(len(tr)):
    ws["B"+str(i+1)]=tr[i]

wb.save("result.xlsx")
wb.close()